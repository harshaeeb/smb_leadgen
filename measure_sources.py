#!/usr/bin/env python3
"""
measure_sources.py — answer one question with evidence instead of a guess:

    Of the businesses Google Places reported as having NO website, how many
    actually have one according to a free, open data source?

This exists because of the Sawyer Brothers case (see CLAUDE.md): Google's
`websiteUri` was simply absent for a business with a real, working site. That
is a data gap, not a bug in leadgen.py, and the only honest way to decide
whether it's worth building a permanent cross-check is to measure how often
it happens on real leads.

Run it against a .xlsx that leadgen.py already produced:

    python measure_sources.py Leads_Plano_TX_plumber_2026-08-18_224406.xlsx

It reads the "No Website" rows, looks each business up in OpenStreetMap via
the Overpass API, and prints how many of them OSM knows a website for.

    # also check Overture Maps (needs: pip install duckdb)
    python measure_sources.py leads.xlsx --overture

THIS IS A MEASUREMENT TOOL, NOT PART OF THE PIPELINE. It imports nothing from
leadgen.py and changes nothing about how leads are qualified. If the numbers
say a cross-check is worth building, that's a separate conversation and a
separate change. If they say it isn't, delete this file.

No API key. Costs nothing. Safe to point at anything.
"""

import argparse
import json
import re
import sys
import time
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

# Nominatim's usage policy requires a genuine identifying User-Agent and no
# more than one request per second. We make exactly one geocode call per run.
USER_AGENT = "smb-leadgen-measure/1.0 (one-off lead-source coverage measurement)"

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
DEFAULT_OVERPASS_URL = "https://overpass-api.de/api/interpreter"

# Search radius around the city centre. Google's own results routinely include
# businesses in neighbouring suburbs (that's how Sawyer Brothers, a Richardson
# business, showed up under a Plano search), so this has to be generous.
DEFAULT_RADIUS_M = 40000

# Words that carry no identifying signal and only hurt name matching.
NAME_NOISE = {
    "llc", "inc", "co", "corp", "company", "the", "and", "of", "a",
    "service", "services", "svc", "plumbing", "plumber", "plumbers",
    "heating", "cooling", "hvac", "electric", "electrical", "electrician",
    "roofing", "landscaping", "lawn", "care", "repair", "solutions",
    "professional", "quality", "best", "local",
}

# OSM tags that can hold a business's website.
OSM_WEBSITE_TAGS = ("website", "contact:website", "url", "contact:url")

SOCIAL_HINTS = ("facebook.com", "instagram.com", "linktr.ee", "m.me", "fb.com")


def _tokens(business_name):
    # Length > 1 drops the "s" that possessives leave behind ("Al's" -> al, s).
    return [w for w in re.findall(r"[a-z0-9]+", business_name.lower()) if len(w) > 1]


def significant_words(business_name):
    """Distinctive words from a business name, for fuzzy matching.

    "Sawyer Brothers Plumber's Service Co" -> ["sawyer", "brothers"]

    Trade words are stripped deliberately: matching on "plumbing" would hit
    every plumber in the county. But when stripping leaves nothing (e.g.
    "Al's Plumbing"), keeping the trade word is far safer than falling back
    to junk tokens -- "al.*plumbing" is a real constraint, "al.*s" matches
    anything. This measurement must under-count rather than over-count: a
    missed match understates the case for building a cross-check, while a
    false match argues for building something unnecessary.
    """
    toks = _tokens(business_name)
    strong = [w for w in toks if w not in NAME_NOISE and len(w) > 2]
    return strong or toks


def is_weak_match(business_name):
    """True when the name has no distinctive words and matching is unreliable."""
    toks = _tokens(business_name)
    return not [w for w in toks if w not in NAME_NOISE and len(w) > 2]


def name_regex(business_name, max_words=2):
    """Overpass-safe case-insensitive regex from a business's distinctive words."""
    words = significant_words(business_name)[:max_words]
    if not words:
        return None
    return ".*".join(re.escape(w) for w in words)


def geocode_city(city, state, session):
    """One Nominatim call per run, to anchor the Overpass radius search."""
    resp = session.get(
        NOMINATIM_URL,
        params={"q": f"{city}, {state}, USA", "format": "json", "limit": 1},
        headers={"User-Agent": USER_AGENT},
        timeout=20,
    )
    resp.raise_for_status()
    hits = resp.json()
    if not hits:
        raise SystemExit(f"Could not geocode '{city}, {state}'. Pass --lat/--lon manually.")
    return float(hits[0]["lat"]), float(hits[0]["lon"])


def normalize_phone(value):
    """Last 10 digits, so '(972) 203-6550' and '+1-972-203-6550' compare equal."""
    digits = re.sub(r"\D", "", value or "")
    return digits[-10:] if len(digits) >= 10 else ""


def osm_lookup(business_name, lat, lon, radius_m, overpass_url, session, expected_phone=None):
    """Look one business up in OSM. Returns a result dict, never raises for
    a miss -- a miss is data, not an error.

    When `expected_phone` is given, an entry whose phone matches is preferred
    over any name-only match and comes back with phone_match=True. Callers
    that act on the result (rather than just counting it) should require that
    flag: name similarity alone pairs up different businesses, and in the
    qualification pipeline a wrong pairing silently deletes a real lead.
    """
    pattern = name_regex(business_name)
    if not pattern:
        return {"found": False, "note": "no distinctive words to match on"}

    # nwr = nodes, ways and relations. Businesses are tagged on all three.
    query = f"""
    [out:json][timeout:30];
    nwr(around:{radius_m},{lat},{lon})["name"~"{pattern}",i];
    out tags 20;
    """
    try:
        resp = session.post(
            overpass_url,
            data={"data": query},
            headers={"User-Agent": USER_AGENT},
            timeout=60,
        )
    except requests.exceptions.RequestException as e:
        return {"found": False, "note": f"overpass request failed ({type(e).__name__})"}

    if resp.status_code != 200:
        return {"found": False, "note": f"overpass HTTP {resp.status_code}"}

    try:
        elements = resp.json().get("elements", [])
    except json.JSONDecodeError:
        return {"found": False, "note": "overpass returned non-JSON (likely rate-limited)"}

    if not elements:
        return {"found": False, "note": "no OSM entry matching that name nearby"}

    want_phone = normalize_phone(expected_phone)
    best = None
    for el in elements:
        tags = el.get("tags", {})
        site = next((tags[t] for t in OSM_WEBSITE_TAGS if tags.get(t)), None)
        osm_phone = tags.get("phone") or tags.get("contact:phone")
        phone_match = bool(want_phone) and normalize_phone(osm_phone) == want_phone
        candidate = {
            "found": True,
            "osm_name": tags.get("name", ""),
            "website": site,
            "phone": osm_phone,
            "phone_match": phone_match,
        }
        # A phone match settles identity, so it wins outright.
        if phone_match:
            return candidate
        # Otherwise prefer an entry that at least carries a website.
        if site and not (best and best.get("website")):
            best = candidate
        best = best or candidate
    return best


def overture_lookup(business_names, lat, lon, release, radius_deg=0.5):
    """Query Overture Places for the whole batch at once via DuckDB.

    Optional: only runs with --overture, and only if duckdb is installed. The
    point of this script is to decide whether a dependency is worth taking, so
    it must not require that dependency to produce a useful answer.
    """
    try:
        import duckdb
    except ImportError:
        print("\n  [overture] skipped: duckdb not installed (pip install duckdb)")
        return {}

    path = (
        f"s3://overturemaps-us-west-2/release/{release}/theme=places/type=place/*"
    )
    con = duckdb.connect()
    con.execute("INSTALL httpfs; LOAD httpfs; SET s3_region='us-west-2';")

    sql = f"""
        SELECT names.primary AS name,
               websites,
               phones
        FROM read_parquet('{path}', filename=true, hive_partitioning=1)
        WHERE bbox.xmin BETWEEN {lon - radius_deg} AND {lon + radius_deg}
          AND bbox.ymin BETWEEN {lat - radius_deg} AND {lat + radius_deg}
    """
    print(f"  [overture] scanning release {release} around ({lat:.3f}, {lon:.3f}) ...")
    try:
        rows = con.execute(sql).fetchall()
    except Exception as e:
        print(f"  [overture] query failed: {str(e)[:200]}")
        print("  [overture] if this is a 404, the release string is wrong -- find the")
        print("             current one at https://docs.overturemaps.org/release/ and")
        print("             pass it with --overture-release")
        return {}

    # Index Overture's names by their distinctive words for fuzzy matching.
    index = []
    for name, websites, phones in rows:
        if not name:
            continue
        index.append((set(significant_words(name)), name, websites, phones))

    results = {}
    for biz in business_names:
        want = set(significant_words(biz)[:2])
        if not want:
            continue
        for words, name, websites, _phones in index:
            if want and want.issubset(words):
                site = websites[0] if websites else None
                results[biz] = {"found": True, "overture_name": name, "website": site}
                break
    return results


def read_no_website_leads(xlsx_path):
    """Pull the 'No Website' rows -- the exact case this is measuring."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Lead Tracker"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))

    def col(name):
        return headers.index(name) if name in headers else None

    i_name, i_tier = col("Business Name"), col("Digital Presence Tier")
    i_addr, i_phone = col("Address"), col("Phone / Contact")
    if i_name is None or i_tier is None:
        raise SystemExit("That doesn't look like a leadgen.py workbook (missing columns).")

    leads = []
    for row in rows:
        if not row or not row[i_name]:
            continue
        if (row[i_tier] or "").strip() != "No Website":
            continue
        leads.append(
            {
                "name": row[i_name],
                "address": row[i_addr] if i_addr is not None else "",
                "phone": row[i_phone] if i_phone is not None else "",
            }
        )
    return leads


def infer_city_state(xlsx_path):
    """The Search Summary sheet records 'Search run: <City>, <ST> — <date>'."""
    wb = load_workbook(xlsx_path, read_only=True)
    if "Search Summary" not in wb.sheetnames:
        return None, None
    first = wb["Search Summary"]["A1"].value or ""
    m = re.search(r"Search run:\s*(.+?),\s*([A-Za-z]{2})\b", str(first))
    return (m.group(1).strip(), m.group(2).strip()) if m else (None, None)


def main():
    ap = argparse.ArgumentParser(
        description="Measure how many 'No Website' leads actually have a website "
        "according to free open data sources. Read-only, no API key, no cost."
    )
    ap.add_argument("xlsx", help="a .xlsx produced by leadgen.py")
    ap.add_argument("--city", default=None, help="override the city (else read from the sheet)")
    ap.add_argument("--state", default=None, help="override the state")
    ap.add_argument("--radius-km", type=float, default=DEFAULT_RADIUS_M / 1000)
    ap.add_argument("--overpass-url", default=DEFAULT_OVERPASS_URL)
    ap.add_argument("--overture", action="store_true", help="also check Overture Maps (needs duckdb)")
    ap.add_argument("--overture-release", default="2026-07-23.0", help="Overture release string")
    ap.add_argument("--limit", type=int, default=None, help="only check the first N leads")
    ap.add_argument("--sleep", type=float, default=2.0, help="seconds between Overpass calls")
    args = ap.parse_args()

    leads = read_no_website_leads(args.xlsx)
    if args.limit:
        leads = leads[: args.limit]
    if not leads:
        raise SystemExit("No 'No Website' leads in that file -- nothing to measure.")

    city, state = args.city, args.state
    if not city or not state:
        inferred_city, inferred_state = infer_city_state(args.xlsx)
        city, state = city or inferred_city, state or inferred_state
    if not city or not state:
        raise SystemExit("Could not determine city/state. Pass --city and --state.")

    print(f"Measuring {len(leads)} 'No Website' leads from {city}, {state}\n")

    session = requests.Session()
    lat, lon = geocode_city(city, state, session)
    print(f"Anchor: {lat:.4f}, {lon:.4f}  (radius {args.radius_km:.0f} km)\n")

    overture_results = {}
    if args.overture:
        overture_results = overture_lookup(
            [l["name"] for l in leads], lat, lon, args.overture_release
        )
        print()

    radius_m = int(args.radius_km * 1000)
    osm_hits = osm_with_site = overture_with_site = 0
    weak_names = []

    for i, lead in enumerate(leads, 1):
        print(f"{i}. {lead['name']}")
        if lead["address"]:
            print(f"   listed address: {lead['address']}")

        weak = is_weak_match(lead["name"])
        if weak:
            weak_names.append(lead["name"])
            print("   (!) generic name -- match unreliable, excluded from the count")

        res = osm_lookup(lead["name"], lat, lon, radius_m, args.overpass_url, session)
        if res.get("found"):
            osm_hits += 1
            site = res.get("website")
            if site:
                if not weak:
                    osm_with_site += 1
                flag = " <-- GOOGLE MISSED THIS"
                if any(s in urlparse(site).netloc.lower() for s in SOCIAL_HINTS):
                    flag = " (social only)"
                elif weak:
                    flag = " (NOT COUNTED - generic name)"
                print(f"   OSM: matched '{res['osm_name']}' -> {site}{flag}")
            else:
                print(f"   OSM: matched '{res['osm_name']}' but no website tag")
        else:
            print(f"   OSM: {res.get('note')}")

        ov = overture_results.get(lead["name"])
        if ov:
            site = ov.get("website")
            if site:
                if not weak:
                    overture_with_site += 1
                print(f"   Overture: matched '{ov['overture_name']}' -> {site} <-- GOOGLE MISSED THIS")
            else:
                print(f"   Overture: matched '{ov['overture_name']}' but no website")
        elif args.overture:
            print("   Overture: no match")

        print()
        if i < len(leads):
            time.sleep(args.sleep)  # be a good Overpass citizen

    n = len(leads)
    countable = n - len(weak_names)
    print("=" * 62)
    print(f"RESULT — of {n} businesses Google reported as having no website:")
    print(f"  OSM had a matching entry:              {osm_hits}/{n}")
    print(f"  OSM knew a website Google missed:      {osm_with_site}/{countable}")
    if args.overture:
        print(f"  Overture knew a website Google missed: {overture_with_site}/{countable}")
    if weak_names:
        print(f"\n  {len(weak_names)} excluded as too generic to match safely:")
        for wn in weak_names:
            print(f"    - {wn}")
    print("=" * 62)

    best = max(osm_with_site, overture_with_site)
    print()
    if countable == 0:
        print("VERDICT: nothing measurable -- every name was too generic.")
        print("         Try a batch with more distinctive business names.")
    elif best == 0:
        print("VERDICT: no evidence a cross-check would help. Don't build it.")
    elif best <= max(1, countable // 10):
        print("VERDICT: marginal. Probably not worth a permanent dependency;")
        print("         the Address column already flags these for manual checking.")
    else:
        print(f"VERDICT: {best} of {countable} checkable 'No Website' leads are wrong.")
        print("         That's a real false-positive rate -- worth building a cross-check.")
    print("\nSpot-check the matches by hand before trusting this number; name")
    print("matching is fuzzy and can pair up two different businesses.")


if __name__ == "__main__":
    main()
