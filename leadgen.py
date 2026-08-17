#!/usr/bin/env python3
"""
leadgen.py — pull local service-business leads with no real professional web
presence for a city/state/service, ranked into a Lead Tracker-formatted
spreadsheet. Built for pitching a professional website + SEO + AI-search
visibility + online booking (cal.com) package.

Usage:
    python leadgen.py "Garland" "TX" "plumber"
    python leadgen.py "Garland" "TX" "dog walker,pet groomer"
    python leadgen.py "Garland" "TX" --industry petcare
    python leadgen.py "Garland" "TX" "electrician" --limit 25 --output electricians.xlsx

Requires:
    GOOGLE_PLACES_API_KEY environment variable (required)

See README.md for setup instructions (getting a key, billing, install steps).

Design notes (why it works this way):
- Google Places Text Search (New) is the sole data source. It's the only
  practical API here that returns a business's actual website field
  (`websiteUri`), which everything else in this script depends on.
- A lead qualifies if it has NO website, only a social media page (Facebook/
  Instagram/etc.), an unreachable/broken website, or a website that fails
  one or more lightweight "is this professional" checks (HTTPS, free
  page-builder subdomain, mobile viewport tag, thin content). A business
  whose site passes all of those checks is not a lead — it already has a
  reasonably professional site and isn't a first target.
- Leads are grouped into two priority tiers: Tier 1 (no site / social-only /
  broken site — effectively zero real web presence) always outranks Tier 2
  (has a site, but it's weak/unprofessional). Within Tier 1, leads are
  ranked by review count then rating, same as before. Within Tier 2, leads
  are ranked by how many quality issues were found (worst sites first),
  falling back to review count then rating.
- A site we were *blocked* from reading is NOT the same as a site that's
  broken, and conflating them is the most expensive mistake this script can
  make. A 403 from a WAF or bot manager means we got turned away; the
  business's website may be perfectly fine. Calling them to say "your website
  is down" would be wrong and would burn the pitch. So anything ambiguous
  (403, 429, 5xx, a bot-challenge page) is held out of the ranked tiers
  entirely and listed on a separate "Could Not Verify" sheet for the operator
  to eyeball. Only failures we can attribute to the site itself — DNS or
  connection failure, TLS failure, timeout, 404, 410 — count as broken and
  earn a Tier 1 spot. Bias every new ambiguous case toward "could not
  verify": missing a lead is cheap, the bad phone call is not.
- AI-search readiness (schema.org/LocalBusiness structured data) is checked
  and reported as an informational column only — it does not affect tier or
  ranking, since it's meant for pitch prep, not lead qualification.
- Website-quality checks are done with plain requests + regex heuristics on
  purpose, no HTML-parsing dependency and no paid auditing API (see
  CLAUDE.md for the cost/quality tradeoff and when to revisit that).
- One run = one batch, capped at --limit (default 25) leads *after* ranking,
  so you're always working the strongest prospects first instead of just
  whatever came back in result order.
- Owner name and email are intentionally left blank. Those come from TDLR /
  TSBPE / county DBA / SOSDirect / Facebook lookups per the enrichment steps
  in Dallas_No_Website_Lead_Generation_Strategy.md — none of those sources
  have a reliable public API, so that step stays manual for now.
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import date
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

GOOGLE_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "")

SOCIAL_DOMAINS = ["facebook.com", "instagram.com", "linktr.ee", "m.me", "fb.com"]

# Free/starter page-builder domains -- a business on one of these (as their
# actual domain, not a custom domain) almost certainly never invested in a
# real site. Some of these platforms (e.g. wordpress.com) also power serious
# self-hosted sites on custom domains, so this only matches the free
# subdomain form, not WordPress in general.
WEAK_BUILDER_DOMAINS = [
    "wixsite.com",
    "weebly.com",
    "godaddysites.com",
    "square.site",
    "blogspot.com",
    "sites.google.com",
    "carrd.co",
    "yolasite.com",
    "webs.com",
    "jimdosite.com",
    "strikingly.com",
    "webnode.com",
    "wordpress.com",
]

# Below this word count on the fetched homepage, treat the site as
# thin/placeholder content rather than a real business site. Heuristic, not
# exact -- see CLAUDE.md.
THIN_CONTENT_WORD_THRESHOLD = 150

REQUEST_TIMEOUT = 8
# Identifies itself honestly as a bot rather than spoofing a browser --
# consistent with this project's stance on not impersonating a browser
# against sites we visit. The cost of being honest is that more sites
# challenge us than would challenge a real browser, which is exactly why the
# "could not verify" bucket exists.
USER_AGENT = "Mozilla/5.0 (compatible; SMBLeadgenBot/1.0)"

# Fingerprints for site builders that render their content with JavaScript.
# We don't run JS, so these pages come back as a near-empty shell and trip the
# thin-content check regardless of how good the real page is. Used to warn
# rather than to qualify -- see detect_js_builder().
JS_BUILDER_MARKERS = (
    ("Wix", ("wixstatic.com", "wix-warmup-data", "parastorage.com", "X-Wix-")),
    ("Squarespace", ("static1.squarespace.com", "squarespace-cdn.com", "SQUARESPACE_CONTEXT")),
    ("Duda", ("irp.cdn-website.com", "dudamobile.com", "d.la4-c1-was.salesforceliveagent")),
    ("Webflow", ("assets.website-files.com", "uploads-ssl.webflow.com", "webflow.js")),
    ("GoDaddy Website Builder", ("img1.wsimg.com", "godaddysites.com")),
    ("Weebly", ("editmysite.com", "weeblysite.com")),
    ("Shopify", ("cdn.shopify.com",)),
)

# Text meaning "a bot wall answered", not "this business has no real site".
# Matched against the start of the response body, lowercased.
CHALLENGE_MARKERS = (
    "just a moment",
    "checking your browser",
    "attention required",
    "cf-browser-verification",
    "enable javascript and cookies",
    "verifying you are human",
    "please verify you are a human",
    "access denied",
    "request unsuccessful",
)

HEADERS = [
    "Rank",
    "Business Name",
    "Niche/Category",
    "# Reviews",
    "Rating",
    "Phone / Contact",
    "Source (Maps/Drive-by)",
    "Digital Presence Tier",
    "Website URL",
    "Website Issues Found",
    "AI Search Ready (schema.org)",
    "Owner Name",
    "Email",
    "Owner/Email Source",
    "Outreach Date",
    "Status",
    "Next Follow-up",
    "Notes",
]

COLUMN_WIDTHS = [6, 30, 20, 10, 8, 16, 18, 24, 32, 40, 16, 20, 26, 20, 14, 12, 14, 30]

# The "Could Not Verify" sheet -- businesses whose site we were blocked from
# reading. Not leads until a human looks, so they get their own tab rather
# than a rank in the call list.
UNVERIFIED_HEADERS = [
    "Business Name",
    "Niche/Category",
    "# Reviews",
    "Rating",
    "Phone / Contact",
    "Website URL",
    "Why We Couldn't Check",
]

UNVERIFIED_COLUMN_WIDTHS = [30, 20, 10, 8, 16, 40, 38]


def normalize(name):
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def is_social_domain(domain):
    return any(s in domain for s in SOCIAL_DOMAINS)


def matching_weak_builder_domain(domain):
    for builder_domain in WEAK_BUILDER_DOMAINS:
        if domain == builder_domain or domain.endswith("." + builder_domain):
            return builder_domain
    return None


def has_https(final_url):
    return urlparse(final_url).scheme == "https"


def extract_visible_text(html_text):
    text = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", html_text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def has_viewport_meta(html_text):
    # Tolerates unquoted attribute values (`name=viewport`, valid HTML5) and
    # whitespace around the `=`. The stricter earlier pattern required quotes
    # and missed real mobile-ready pages.
    return bool(re.search(r'(?i)<meta\b[^>]*\bname\s*=\s*["\']?viewport["\'\s>]', html_text))


def detect_js_builder(html_text):
    """Name the JS site builder a page was made with, if we recognise it.

    Purely diagnostic. These platforms render their content client-side, so
    the HTML we fetch is a near-empty shell and `is_thin_content()` will flag
    them no matter how substantial the real page is. Knowing the builder lets
    us say "we can't measure this" instead of quietly reporting a bad signal.
    """
    for name, markers in JS_BUILDER_MARKERS:
        if any(m.lower() in html_text.lower() for m in markers):
            return name
    return None


def is_thin_content(html_text):
    return len(extract_visible_text(html_text).split()) < THIN_CONTENT_WORD_THRESHOLD


def has_schema_markup(html_text):
    if "schema.org" in html_text and re.search(r'(?i)application/ld\+json', html_text):
        return True
    return bool(re.search(r'(?i)itemtype=["\']https?://schema\.org/', html_text))


def looks_like_challenge(html_text):
    head = html_text[:4096].lower()
    return any(marker in head for marker in CHALLENGE_MARKERS)


def classify_response(resp):
    """Decide what an HTTP response tells us before we bother reading the page.

    Returns a verdict dict for anything conclusive, or None when the response
    looks like a real page worth analyzing.

    The split between "broken" and "blocked" is the point of this function.
    "broken" is a claim about the *business's site* and earns a Tier 1 lead,
    so it's reserved for failures we can actually attribute to them. "blocked"
    means we were turned away and learned nothing -- it must never rank.
    """
    status = resp.status_code

    # Cloudflare sets this when it mitigates a request, sometimes alongside a
    # 200, so it has to be checked before the status code.
    if "cf-mitigated" in resp.headers:
        return {"status": "blocked", "reason": f"Bot protection (cf-mitigated, HTTP {status})"}

    if status in (403, 429):
        return {"status": "blocked", "reason": f"Blocked by bot protection (HTTP {status})"}

    if status in (404, 410):
        return {"status": "broken", "reason": f"Homepage returns HTTP {status}"}

    if status >= 500:
        # Could be genuinely down (a strong lead) or a blip. Not worth the
        # risk of being wrong on a cold call.
        return {"status": "blocked", "reason": f"Server error (HTTP {status}) - may be transient"}

    if status >= 400:
        return {"status": "blocked", "reason": f"Unexpected HTTP {status}"}

    return None


def analyze_website(website_url):
    """Fetch a candidate's listed website once and classify it.

    Returns a dict with a "status" of:
      "social_only"  -- Tier 1 lead, no fetch attempted
      "broken"       -- Tier 1 lead; the site genuinely fails for real visitors
      "blocked"      -- NOT a lead; we were turned away, verdict unknown
      "weak"         -- Tier 2 lead; also carries "issues" and "ai_ready"
      "professional" -- not a lead, the site is fine
    "broken" and "blocked" also carry a human-readable "reason".
    """
    domain = urlparse(website_url).netloc.lower()
    if not domain:
        return {"status": "broken", "reason": "Malformed website URL on the listing"}
    if is_social_domain(domain):
        return {"status": "social_only"}

    try:
        resp = requests.get(
            website_url,
            headers={"User-Agent": USER_AGENT},
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
        )
    # Order matters: ConnectTimeout subclasses both Timeout and ConnectionError,
    # and SSLError subclasses ConnectionError.
    except requests.exceptions.Timeout:
        return {"status": "broken", "reason": f"No response within {REQUEST_TIMEOUT}s"}
    except requests.exceptions.SSLError:
        # A bad certificate is broken for every real visitor too -- browsers
        # show a full-page security warning.
        return {"status": "broken", "reason": "TLS/certificate failure"}
    except requests.exceptions.ConnectionError:
        return {"status": "broken", "reason": "Could not connect (DNS or connection failure)"}
    except requests.exceptions.RequestException as e:
        return {"status": "broken", "reason": f"Request failed ({type(e).__name__})"}

    verdict = classify_response(resp)
    if verdict:
        return verdict

    html_text = resp.text or ""
    if looks_like_challenge(html_text):
        return {"status": "blocked", "reason": "Bot challenge page served instead of the site"}

    issues = []
    if not has_https(resp.url):
        issues.append("No HTTPS")
    builder_match = matching_weak_builder_domain(urlparse(resp.url).netloc.lower())
    if builder_match:
        issues.append(f"Free page-builder ({builder_match})")
    if not has_viewport_meta(html_text):
        issues.append("No mobile viewport tag")
    if is_thin_content(html_text):
        issues.append("Thin/minimal content")

    ai_ready = has_schema_markup(html_text)
    detail = {
        "final_url": resp.url,
        "http_status": resp.status_code,
        "word_count": len(extract_visible_text(html_text).split()),
        "js_builder": detect_js_builder(html_text),
        "ai_ready": ai_ready,
    }
    if issues:
        return {"status": "weak", "issues": issues, "ai_ready": ai_ready, "detail": detail}
    return {"status": "professional", "ai_ready": ai_ready, "detail": detail}


def explain(website_url):
    """Show exactly what the checks saw for one URL.

    For answering "why did this business land on my list?" without guessing.
    Prints per-check pass/fail plus the raw numbers behind them.
    """
    print(f"Checking {website_url}\n")
    analysis = analyze_website(website_url)
    status = analysis["status"]

    if status == "social_only":
        print("  Verdict: TIER 1 LEAD - Social Only (no real site)")
        print("  The listed 'website' is a social media page, so no site was fetched.")
        return
    if status in ("broken", "blocked"):
        label = "TIER 1 LEAD - Website Unreachable/Broken" if status == "broken" else (
            "NOT A LEAD - held for manual verification"
        )
        print(f"  Verdict: {label}")
        print(f"  Reason:  {analysis['reason']}")
        if status == "blocked":
            print("\n  We were turned away, so nothing is known about the real site.")
        return

    d = analysis["detail"]
    checks = [
        ("HTTPS", has_https(d["final_url"]), "loads over https://"),
        (
            "Custom domain",
            matching_weak_builder_domain(urlparse(d["final_url"]).netloc.lower()) is None,
            "not on a free page-builder subdomain",
        ),
        ("Mobile viewport tag", "No mobile viewport tag" not in analysis.get("issues", []), "has <meta name=viewport>"),
        (
            "Content volume",
            "Thin/minimal content" not in analysis.get("issues", []),
            f"{d['word_count']} words found, need {THIN_CONTENT_WORD_THRESHOLD}",
        ),
    ]

    if d["final_url"] != website_url:
        print(f"  Redirected to: {d['final_url']}")
    print(f"  HTTP status:   {d['http_status']}\n")

    for name, passed, detail_text in checks:
        mark = "PASS" if passed else "FAIL"
        print(f"    {name:<22} {mark}   ({detail_text})")
    print(f"    {'schema.org markup':<22} {'yes' if d['ai_ready'] else 'no':<6} (informational only, never affects ranking)")

    print()
    if status == "professional":
        print("  Verdict: NOT A LEAD - the site passes every check.")
    else:
        print(f"  Verdict: TIER 2 LEAD - Weak/Unprofessional Website")
        print(f"  Issues:  {', '.join(analysis['issues'])}")

    if d["js_builder"] and "Thin/minimal content" in analysis.get("issues", []):
        print(
            f"\n  ⚠ LIKELY FALSE POSITIVE. This page was built with {d['js_builder']}, which\n"
            f"    renders its content with JavaScript. This script doesn't run JavaScript,\n"
            f"    so it only sees a {d['word_count']}-word shell -- the real page a visitor\n"
            f"    sees is probably much bigger. Treat 'Thin/minimal content' as unmeasured\n"
            f"    here, not as a finding."
        )


def google_text_search(query, api_key, max_pages=3):
    """Text Search (New). websiteUri triggers the Enterprise SKU — see README
    for the current per-1000-request cost and free monthly credit."""
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": (
            "places.displayName,places.nationalPhoneNumber,"
            "places.websiteUri,places.formattedAddress,"
            "places.rating,places.userRatingCount,nextPageToken"
        ),
    }
    results = []
    body = {"textQuery": query}
    for page in range(max_pages):
        resp = requests.post(url, headers=headers, json=body, timeout=20)
        if resp.status_code != 200:
            print(f"    Google Places error {resp.status_code}: {resp.text[:300]}")
            break
        data = resp.json()
        results.extend(data.get("places", []))
        token = data.get("nextPageToken")
        if not token:
            break
        time.sleep(2)  # Google requires a short delay before a page token is valid
        body = {"textQuery": query, "pageToken": token}
    return results


def resolve_categories(service_arg, industry, categories_file):
    """A specific --service (or positional) list always wins. Otherwise fall
    back to a named preset in categories.json (default preset: trades)."""
    if service_arg:
        return [s.strip() for s in service_arg.split(",") if s.strip()]
    with open(categories_file) as f:
        cats = json.load(f)
    if industry not in cats:
        raise SystemExit(
            f"Unknown --industry '{industry}'. Options in {categories_file}: {list(cats)}"
        )
    return cats[industry]


def run(city, state, categories):
    """Returns (leads, unverified, stats).

    `unverified` holds businesses whose site we were blocked from reading.
    They are kept out of `leads` entirely so a site we simply couldn't check
    can never displace a real prospect from the ranked batch.
    """
    location = f"{city}, {state}"
    all_leads = []
    unverified = []
    stats = []
    seen_keys = set()

    for category in categories:
        query = f"{category} in {location}"
        print(f"Searching: {query}")
        places = google_text_search(query, GOOGLE_KEY)
        total = len(places)
        qualifying_count = 0
        unverified_count = 0

        for p in places:
            name = p.get("displayName", {}).get("text", "")
            phone = p.get("nationalPhoneNumber", "")
            website = p.get("websiteUri", "")
            key = phone or normalize(name)
            if not key or key in seen_keys:
                continue
            # Claim the key up front: whatever we decide about this business,
            # a duplicate listing shouldn't trigger a second HTTP fetch.
            seen_keys.add(key)

            reviews = p.get("userRatingCount", 0) or 0
            rating = p.get("rating", 0) or 0

            if not website:
                tier, severity = 0, 0
                presence_label = "No Website"
                website_url, issues_str, ai_ready_str = "", "", "N/A"
            else:
                analysis = analyze_website(website)
                status = analysis["status"]
                if status == "professional":
                    continue  # has a real, working, reasonably modern site -- not a lead

                if status == "blocked":
                    # We learned nothing about this site. Park it for manual
                    # review rather than guessing -- see the module docstring.
                    unverified_count += 1
                    unverified.append(
                        {
                            "Business Name": name,
                            "Niche/Category": category,
                            "# Reviews": reviews,
                            "Rating": rating,
                            "Phone / Contact": phone,
                            "Website URL": website,
                            "Why We Couldn't Check": analysis["reason"],
                        }
                    )
                    continue

                website_url = website
                if status == "social_only":
                    tier, severity = 0, 0
                    presence_label = "Social Only (no real site)"
                    issues_str, ai_ready_str = "", "N/A"
                elif status == "broken":
                    tier, severity = 0, 0
                    presence_label = "Website Unreachable/Broken"
                    issues_str, ai_ready_str = analysis["reason"], "N/A"
                else:  # weak
                    tier = 1
                    severity = len(analysis["issues"])
                    presence_label = "Weak/Unprofessional Website"
                    issues_str = ", ".join(analysis["issues"])
                    ai_ready_str = "Yes" if analysis["ai_ready"] else "No"

            qualifying_count += 1
            all_leads.append(
                {
                    "Business Name": name,
                    "Niche/Category": category,
                    "# Reviews": reviews,
                    "Rating": rating,
                    "Phone / Contact": phone,
                    "Source (Maps/Drive-by)": "Google Maps",
                    "Digital Presence Tier": presence_label,
                    "Website URL": website_url,
                    "Website Issues Found": issues_str,
                    "AI Search Ready (schema.org)": ai_ready_str,
                    "Owner Name": "",
                    "Email": "",
                    "Owner/Email Source": "",
                    "Outreach Date": "",
                    "Status": "New",
                    "Next Follow-up": "",
                    "Notes": "",
                    "_tier": tier,
                    "_severity": severity,
                }
            )

        stats.append((category, total, qualifying_count, unverified_count))
        suffix = f", {unverified_count} could not be verified" if unverified_count else ""
        print(f"  -> {total} results, {qualifying_count} qualifying leads{suffix}")
        time.sleep(1)

    return all_leads, unverified, stats


def rank_and_trim(all_leads, limit):
    tier0 = [l for l in all_leads if l["_tier"] == 0]
    tier1 = [l for l in all_leads if l["_tier"] == 1]

    tier0_sorted = sorted(
        tier0, key=lambda l: (to_num(l.get("# Reviews")), to_num(l.get("Rating"))), reverse=True
    )
    tier1_sorted = sorted(
        tier1,
        key=lambda l: (l["_severity"], to_num(l.get("# Reviews")), to_num(l.get("Rating"))),
        reverse=True,
    )

    ranked = tier0_sorted + tier1_sorted
    kept = ranked[:limit]
    dropped = len(ranked) - len(kept)
    return kept, dropped


def write_output(leads, unverified, stats, dropped, limit, output_path, city, state):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Tracker"

    header_fill = PatternFill("solid", start_color="1F6FEB", end_color="1F6FEB")

    def style_header(sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    ws.append(HEADERS)
    style_header(ws)

    for i, lead in enumerate(leads, start=1):
        row = [i] + [lead.get(h, "") for h in HEADERS[1:]]
        ws.append(row)

    for col_letter, width in zip("ABCDEFGHIJKLMNOPQR", COLUMN_WIDTHS):
        ws.column_dimensions[col_letter].width = width

    # Businesses whose site we couldn't read. Deliberately a separate sheet:
    # the Lead Tracker tab stays a clean 1..N ranked call list, and these
    # need a human to look before they're worth a call.
    unchecked = wb.create_sheet("Could Not Verify")
    unchecked.append(UNVERIFIED_HEADERS)
    style_header(unchecked)
    if unverified:
        for row in sorted(unverified, key=lambda r: to_num(r["# Reviews"]), reverse=True):
            unchecked.append([row.get(h, "") for h in UNVERIFIED_HEADERS])
    else:
        unchecked.append(["No leads needed manual verification in this run."])
    for col_letter, width in zip("ABCDEFG", UNVERIFIED_COLUMN_WIDTHS):
        unchecked.column_dimensions[col_letter].width = width

    summary = wb.create_sheet("Search Summary")
    summary.append([f"Search run: {city}, {state} — {date.today().isoformat()}"])
    summary.append([f"Batch limit: top {limit}, Tier 1 (no real site) ranked above Tier 2 (weak site)"])
    summary.append([f"Candidates found beyond the limit (not included): {dropped}"])
    summary.append([])
    summary.append(["Category", "# Results", "# Qualifying Leads", "# Could Not Verify"])
    for cell in summary[5]:
        cell.font = Font(bold=True)
    for row in stats:
        summary.append(list(row))
    summary.append([])
    summary.append(["Total leads in this batch:", len(leads)])
    summary.append(["Held for manual verification:", len(unverified)])
    summary.column_dimensions["A"].width = 30

    wb.save(output_path)
    print(f"\nSaved {len(leads)} leads (top {limit}, {dropped} more found but not included) to {output_path}")
    if unverified:
        print(
            f"{len(unverified)} business(es) blocked our check and are on the "
            f"'Could Not Verify' tab — open those sites yourself before calling."
        )


def main():
    parser = argparse.ArgumentParser(
        description="Pull a ranked, top-N lead batch (no website or unprofessional website) for a city/state/service into a Lead Tracker sheet."
    )
    parser.add_argument("city", nargs="?", default=None, help='e.g. "Garland"')
    parser.add_argument("state", nargs="?", default=None, help='e.g. "TX"')
    parser.add_argument(
        "service",
        nargs="?",
        default=None,
        help='Service type(s) to search, comma-separated, e.g. "plumber" or "dog walker,pet groomer". '
        "Overrides --industry when given.",
    )
    parser.add_argument(
        "--industry",
        default="trades",
        help="Category preset from categories.json, used only if 'service' is omitted (default: trades). Options: trades, petcare",
    )
    parser.add_argument(
        "--categories-file",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "categories.json"),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=25,
        help="Max leads to keep per batch/run, ranked by tier then review count/rating (default: 25)",
    )
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    parser.add_argument(
        "--explain",
        metavar="URL",
        default=None,
        help="Diagnose one website instead of running a search: shows every check, "
        "pass/fail, and the numbers behind them. Answers 'why did this business land "
        "on my list?'. Needs no API key and costs nothing.",
    )
    args = parser.parse_args()

    # Diagnostic mode: no Places search, so no API key and no spend.
    if args.explain:
        explain(args.explain)
        return

    if not args.city or not args.state:
        parser.error("city and state are required (or use --explain URL to diagnose one site)")

    if not GOOGLE_KEY:
        sys.exit(
            "GOOGLE_PLACES_API_KEY is not set. See README.md for how to get one, "
            "then set it as an environment variable before running this script."
        )

    categories = resolve_categories(args.service, args.industry, args.categories_file)
    label = args.service or args.industry
    output_path = (
        args.output
        or f"Leads_{args.city.replace(' ', '_')}_{args.state}_{label.replace(' ', '_').replace(',', '-')}_{date.today().isoformat()}.xlsx"
    )

    all_leads, unverified, stats = run(args.city, args.state, categories)
    kept, dropped = rank_and_trim(all_leads, args.limit)
    write_output(kept, unverified, stats, dropped, args.limit, output_path, args.city, args.state)


if __name__ == "__main__":
    main()
